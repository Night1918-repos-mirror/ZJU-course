# !/usr/bin/env python
# -*- coding:utf-8 -*-

# Implement AETG algorithm for long'ce website and jing'dong shopping website.
# url_longce: https://prod.dragontesting.com/tool/login.html#/
# url_jingdong: https://list.jd.com/list.html?cat=737,794,798

'''
    Covering Array  CA(N; K, T, V):
        N is the number of test cases,
        K is the number of parameters,
        T is the intensity of this CA.
        In this array, any N*T subarray contains all t-tuples of this t parameters.

    Greedy Algorithm:
        Assume that the system has K parameters, each parameter ki has li possible values.
        1. Choose the parameter i randomly.
        2. Assign value for i randomly.
        3. loop for 1-2, until all parameters are assigned. Generate test case T. 
        4. loop for 1-3 C times, where C is the number of candidates.
        5. select the test case which covers the most uncovered pairs from C candidates, and add it into CA
        6. loop for 1-5 until all pairs are covered. CA generates successfully.
        7. repeat 1-6 for R times, and choose the smallest CA.

'''

import time
import random

import numpy as np
import openpyxl


def test_case_jd():
    header = ['品牌', '屏幕比例', '能效等级', '固态硬盘(SSD)', '机械硬盘', '内存容量', '屏幕尺寸',
              '屏幕色域', '屏幕刷新率', '覆盖数量（不包括已经覆盖的）']
    brand = ['all', 'thinkpad', 'dell', 'huawei', 'hp', 'lenvo', 'apple', 'asus', 'mi', 'honor',
             'dere', 'acer', 'mechrevo', 'microsoft', 'haier']
    size = ['all', '16:10', '16:9', '4:3', '3:2']
    efficiency = ['all', '1', '2', '3']
    ssd = ['all', '3TB', '128GB', '256GB+1TB', '512GB+2TB', '3TB2', '64GB', '240GB', '4TB*2', '256GB', '4TB*3',
           '512GB', '512GB*2', '512GB+1TB', '4TB', '8TB', '2TB+4TB', '无固态硬盘']
    material = ['all', '金属', '金属+复合', '复合材质', '含碳纤维']
    capacity = ['all', '4GB', '6GB', '48GB', '18GB', '192GB', '8GB', '12GB',
                '16GB', '20GB', '24GB', '32GB', '36GB', '40GB', '64GB', '96GB', '128GB']
    refresh_rate = ['all', '60Hz', '165Hz', '360Hz', '60Hz-120Hz', '90Hz', '120Hz', '144Hz', '240Hz']

    params = [16, 5, 4, 21, 5, 17, 9]
    CA, cover_num = AETG(params, 7, 2)

    # write into xlsx.
    wb = openpyxl.Workbook()
    st = wb.create_sheet("jing'dong", index=0)

    # write header
    col_1 = []
    for i in range(len(CA)):
        col_1.append(i + 1)

    for i, item in enumerate(header):
        st.cell(row=1, column=i + 2, value=item)
    for i in range(len(CA)):
        st.cell(row=i + 2, column=1, value='case_' + str(col_1[i]))

    # write data
    for i, test_case in enumerate(CA):
        for j, item in enumerate(test_case):
            _row, _col = i + 2, j + 2
            if j == 0:
                st.cell(row=_row, column=_col, value=brand[item])
            elif j == 1:
                st.cell(row=_row, column=_col, value=size[item])
            elif j == 2:
                st.cell(row=_row, column=_col, value=efficiency[item])
            elif j == 3:
                st.cell(row=_row, column=_col, value=ssd[item])
            elif j == 4:
                st.cell(row=_row, column=_col, value=material[item])
            elif j == 5:
                st.cell(row=_row, column=_col, value=capacity[item])
            elif j == 6:
                st.cell(row=_row, column=_col, value=refresh_rate[item])
        # write cover number
        st.cell(row=i + 2, column=9, value=str(cover_num[i]))

    # save
    this_time = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
    file_name = 'jingdong ' + '.xlsx'
    wb.save(file_name)


def test_case_xc():
    header = ['出发地', '目的地', '出发日期', '返回日期', '乘客数量', '直飞', '舱型', '覆盖数量（不包括已经覆盖的）']
    departure = ['beijing', 'shanghai', 'guangzhou', 'shenzhen', 'chengdu', 'chongqing', 'hangzhou', 'nanjing']
    destination = ['beijing', 'shanghai', 'guangzhou', 'shenzhen', 'chengdu', 'chongqing', 'hangzhou', 'nanjing']
    departure_date = ['2021-05-01', '2021-05-02', '2021-05-03', '2021-05-04', '2021-05-05', '2021-05-06']
    return_date = ['2021-05-01', '2021-05-02', '2021-05-03', '2021-05-04', '2021-05-05', '2021-05-06']
    passenger = ['1', '2', '3', '4', '5', '5', '6', '7', '8', '9', '10', '11', '12', '13']
    direct_flight = ['true', 'false']
    cabin_type = ['经济/超经舱', '公务/头等舱', '公务舱', '头等舱']

    params = [8, 8, 6, 6, 14, 2, 4]
    CA, cover_num = AETG(params, 7, 2)

    # write into xlsx.
    wb = openpyxl.Workbook()
    st = wb.create_sheet("xie'cheng", index=0)

    # write header
    col_1 = []
    for i in range(len(CA)):
        col_1.append(i + 1)

    for i, item in enumerate(header):
        st.cell(row=1, column=i + 2, value=item)
    for i in range(len(CA)):
        st.cell(row=i + 2, column=1, value='case_' + str(col_1[i]))

    # write data
    for i, test_case in enumerate(CA):
        for j, item in enumerate(test_case):
            _row, _col = i + 2, j + 2
            if j == 0:
                st.cell(row=_row, column=_col, value=departure[item])
            elif j == 1:
                st.cell(row=_row, column=_col, value=destination[item])
            elif j == 2:
                st.cell(row=_row, column=_col, value=departure_date[item])
            elif j == 3:
                st.cell(row=_row, column=_col, value=return_date[item])
            elif j == 4:
                st.cell(row=_row, column=_col, value=passenger[item])
            elif j == 5:
                st.cell(row=_row, column=_col, value=direct_flight[item])
            elif j == 6:
                st.cell(row=_row, column=_col, value=cabin_type[item])
        # write cover number
        st.cell(row=i + 2, column=9, value=str(cover_num[i]))

    # save
    this_time = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
    file_name = 'xiecheng ' + '.xlsx'
    wb.save(file_name)


def AETG(params: list, _params_num, wise_num):
    """
    Implementation for AETG algorithm.
    :description:
        AETG Algorithm:
        Assume that the system has K parameters, each parameter ki has li possible values.
        when we have already selected r test cases into CA and go to the r+1 one,
        1. Have chose a number M as the candidates number.
        2. Choose a parameter fi and value vj for fi, where vj has a highest frequency in the uncovered pairs.
        3. Select fi as f1, and make a random permutation for remaining parameters.
        4. Assume that f1-fj is assigned, then assign vj+1 for fj+1 that vj+1 has the most number of uncovered pair
            with all assigned parameters.
        5. loop for 1-4 to get M candidates and find the optimal one, add into CA.
        6. loop for 1-5 until all pairs are covered. CA generates successfully.
        7. repeat 1-6 for R times, and choose the smallest CA.

    :param params:      LIST    list contains the number of possible value of each params
    :param _params_num: INT     the number of parameters.
    :param wise_num:    INT     wise number, which represents the intensity of CA.
    :return:            MATRIX  represents CA, each position is a index of value as to a specified parameter.
    """

    # params is the list contains ki
    params_num = len(params)  # K
    assert (params_num == _params_num)

    covered_pairs_data = set()  # contains all pairs(including higher dimension)

    uncovered_pairs_data = set()  # contains all uncovered pairs(including higher dimension)

    all_pairs_num_data = calc_covered_pairs_num(params, uncovered_pairs_data, wise_num)

    M = 50  # M is the number of candidates. STEP 1

    R = 10  # repeat the algorithm for R times

    candidate_CA = []

    for i in range(R):  # STEP 7

        # using data to avoid repeat calculation
        covered_pairs = set()

        uncovered_pairs = uncovered_pairs_data.copy()

        all_pairs_num = all_pairs_num_data

        CA = []  # the matrix of Covering Array, which is the return value

        while len(covered_pairs) != all_pairs_num:  # STEP 6
            candidates = []
            new_case_each_candidate_generate = []
            for j in range(M):  # STEP 5

                candidate = [-1 for _ in range(params_num)]  # list with length at param_num to contain a candidate
                has_selected_params = set()

                # STEP 2
                # find (f1, v1), where v1 has the highest frequency in uncovered pairs.
                uncovered_pairs_list = list(uncovered_pairs)
                # print(uncovered_pairs_list)
                max_cnt = -1
                f1, v1 = 0, 0
                for k, param in enumerate(params):  # k, the index of param
                    # param, the number of possible value of param k
                    for l in range(param):  # l, index of each possible value
                        # eg: k = 0, param = 2, l = 1
                        cnt = 0  # cnt, the number of uncovered pairs with l
                        for uncovered_pair in uncovered_pairs_list:
                            if uncovered_pair[k] == l:
                                cnt += 1
                        if cnt > max_cnt:
                            f1, v1 = k, l
                            max_cnt = cnt
                candidate[f1] = v1
                has_selected_params.add(f1)
                has_selected_params_num = len(has_selected_params)

                # STEP 3, 4
                while (True):
                    # randomly choose the next parameter, which is not selected yet.
                    while (True):
                        f = random.randint(0, params_num - 1)  # if we have 5 params, then it produces
                        # next param index randomly from 0 to 4

                        if f not in has_selected_params:  # if this param has not been selected, break
                            break  # otherwise, loop the production
                    v = 0

                    if has_selected_params_num < wise_num:
                        # already selected parameter number < wise_num(t)
                        # we need to choose next value, combined it with all selected parameter
                        # let this combination have a highest frequency in uncovered pairs
                        max_frequency = -1
                        for vi in range(params[f]):
                            frequency = 0
                            pairs_combined_selected_param_and_next_value = candidate.copy()
                            pairs_combined_selected_param_and_next_value[f] = vi
                            for uncovered_pair_tuple in uncovered_pairs_list:  # TUPLE
                                uncovered_pair = list(uncovered_pair_tuple)
                                if is_part_of_uncovered_pair(uncovered_pair,
                                                             pairs_combined_selected_param_and_next_value):
                                    frequency += 1
                            if frequency > max_frequency:
                                v = vi
                                max_frequency = frequency

                    else:
                        # already selected parameter number(k) >= wise_num(t)
                        # we need to calculate C(k, t-1) times, where C is the combination number of k and t-1
                        # each time choose t-1 parameters from the k selected parameters,
                        # and combined them with next value. Find the uncovered pairs they covered.
                        # let the number have a highest frequency.
                        max_frequency = -1
                        for vi in range(params[f]):
                            current_selected_params = candidate.copy()
                            current_selected_params_num = has_selected_params_num
                            frequency = calc_covered_ucp_num_for_next_value(
                                current_selected_params, current_selected_params_num, wise_num, f, vi,
                                uncovered_pairs)
                            if frequency > max_frequency:
                                v = vi
                                max_frequency = frequency

                    # update
                    candidate[f] = v
                    has_selected_params.add(f)
                    has_selected_params_num = len(has_selected_params)

                    # after all parameters are selected, the test case is fully generated.
                    if has_selected_params_num == params_num:
                        break

                # after all parameters are selected, add it into candidates
                candidates.append(candidate)
                print("the new candidate " + str(len(candidates)) + " is " + str(candidate))

            # calculate new case each candidate gernerate
            for candidate in candidates:
                new_case_generated = calc_new_case_generated(uncovered_pairs, candidate, wise_num)
                new_case_each_candidate_generate.append(new_case_generated)

            optimal_candidate_index = new_case_each_candidate_generate.index(max(new_case_each_candidate_generate))
            update_covered_pairs(uncovered_pairs, covered_pairs, candidates[optimal_candidate_index], wise_num)
            CA.append(candidates[optimal_candidate_index])
            print("the test case " + str(len(CA)) + " is " + str(candidates[optimal_candidate_index]))
            # find the test case that generate the most number of new pairs and append it.
        candidate_CA.append(CA)
        # print(CA)

    _ret_CA = get_optimal_CA(candidate_CA)  # find the optimal CA which has the least number of N.

    # calculate each test case covers how many pairs
    cover_num = []
    ret_CA = []
    for test_case in _ret_CA:
        subset_num = 2 ** params_num
        total_cover_num_for_one_case = 0
        for i in range(subset_num):
            subset = []
            subset_len = 0
            for j in range(params_num):
                if ((i >> j) % 2):
                    subset.append(test_case[j])
                    subset_len += 1
                else:
                    subset.append(-1)
            if subset_len == wise_num:
                if tuple(subset) in uncovered_pairs_data:
                    total_cover_num_for_one_case += 1
                    uncovered_pairs_data.remove(tuple(subset))
        if total_cover_num_for_one_case != 0:
            ret_CA.append(test_case)
            cover_num.append(total_cover_num_for_one_case)
    return ret_CA, cover_num


def calc_covered_pairs_num(param_list, pairs, wise_num):
    """
    Calculate the number of covered pairs in a intensity of wise_num and get all pairs.
    :description:
        traversal the set's all subsets where the elements number is wise_num.
        the subset represents that choosing wise_num parameters, each of them contains how many possible value.
        The possible combination is thr product of each number.

    :param  set:      LIST   list contains the number of possible value of each params
    :param  wise_num: INT    wise number, which represents the intensity of CA.
    :param  pairs     SET    set to contain all uncovered pairs. Get them in this function.
    :return:total_num INT    the number of covered_pairs.
    """

    total_num = 0

    # param_list = [2,2,4,4,6]

    elements_num = len(param_list)
    subset_num = 2 ** elements_num  # number of subsets for a set with n elements is 2^n

    # using binary string to traversal all subsets.
    for i in range(subset_num):
        subset = []
        subset_nonegative_num = 0
        for j in range(elements_num):
            if (i >> j) % 2:
                subset.append(param_list[j])
                subset_nonegative_num += 1  # length exclude -1
            else:
                subset.append(-1)
        # if len(subset) == wise_num:     # if the elements number == wise_num
        #     total_num += np.prod(subset)

        if subset_nonegative_num == wise_num:  # specified wise_num
            total_num += (int)(np.prod(subset) / ((-1) ** (len(subset) - subset_nonegative_num)))
            # print(subset)
            get_uncovered_pairs_in_subset(subset, 0, pairs, [])
    print("the total uncovered pairs number is " + str(total_num))
    # print(pairs)
    return total_num


def get_optimal_CA(candidates):
    N_of_each_CA = []
    for candidate in candidates:
        N_of_each_CA.append(len(candidate))
    return candidates[N_of_each_CA.index(min(N_of_each_CA))]


def get_uncovered_pairs_in_subset(subset, idx, pairs: set, pair: list):
    if idx + 1 == len(subset):
        if subset[idx] > 0:
            for i in range(subset[idx]):
                pairs.add(tuple(pair + [i]))
        else:
            pairs.add(tuple(pair + [-1]))
        return
    if subset[idx] > 0:
        for i in range(subset[idx]):
            get_uncovered_pairs_in_subset(subset, idx + 1, pairs, pair + [i])
    else:
        get_uncovered_pairs_in_subset(subset, idx + 1, pairs, pair + [-1])


def is_part_of_uncovered_pair(uncovered_pair, part_of_uncovered_pair):
    assert len(uncovered_pair) == len(part_of_uncovered_pair)
    length = len(uncovered_pair)
    for i in range(length):
        if uncovered_pair[i] == -1 and part_of_uncovered_pair[i] == -1:
            continue
        elif uncovered_pair[i] == part_of_uncovered_pair[i]:
            continue
        elif part_of_uncovered_pair[i] != -1:
            return False
    return True


def calc_covered_ucp_num_for_next_value(selected_params, selected_params_num, wise_num, f, vi, uncovered_pairs):
    elem_num = len(selected_params)
    subset_num = 2 ** elem_num
    total_cover_ucp_num = 0
    for i in range(subset_num):
        subset = []
        subset_len = 0
        for j in range(elem_num):
            if ((i >> j) % 2):
                subset.append(selected_params[j])
                subset_len += 1
            else:
                subset.append(-1)
        if subset_len == wise_num - 1:
            subset[f] = vi
            if tuple(subset) in uncovered_pairs:
                total_cover_ucp_num += 1
    return total_cover_ucp_num


def calc_new_case_generated(uncovered_pairs, candidate: list, wise_num):
    elem_num = len(candidate)
    subset_num = 2 ** elem_num
    total_new_pairs_num = 0
    for i in range(subset_num):
        subset = []
        subset_len = 0
        for j in range(elem_num):
            if (i >> j) % 2:
                subset.append(candidate[j])
                subset_len += 1
            else:
                subset.append(-1)
        if subset_len == wise_num:
            if tuple(subset) in uncovered_pairs:
                total_new_pairs_num += 1
    return total_new_pairs_num


def update_covered_pairs(uncovered_pairs, covered_pairs, candidate, wise_num):
    elem_num = len(candidate)
    subset_num = 2 ** elem_num
    for i in range(subset_num):
        subset = []
        subset_len = 0
        for j in range(elem_num):
            if (i >> j) % 2:
                subset.append(candidate[j])
                subset_len += 1
            else:
                subset.append(-1)
        if subset_len == wise_num:
            if tuple(subset) in uncovered_pairs:
                uncovered_pairs.remove(tuple(subset))
                covered_pairs.add(tuple(subset))


if __name__ == '__main__':
    test_case_jd()
    # test_case_xc()
